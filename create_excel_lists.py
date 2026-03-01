"""
Creates four Excel reference files for Em to review and edit:
1. Weekly tasks (exhaustive, all 40 weeks)
2. Hospital bag checklist
3. Baby essentials checklist
4. Work & admin checklist
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SAGE    = "7a9e7e"
CLAY    = "c4956a"
WHITE   = "FFFFFF"
LIGHT   = "f5f0e8"
HEADER  = "4a3f35"


def style_header_row(ws, row, cols, bg=SAGE, fg=WHITE, size=11, bold=True):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=fg, size=size, bold=bold)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_subheader_row(ws, row, cols, bg=CLAY, fg=WHITE):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=fg, size=10, bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_data_row(ws, row, cols, bg=None):
    fill = PatternFill("solid", fgColor=bg) if bg else PatternFill("none")
    font = Font(size=10)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if bg:
            cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def thin_border():
    thin = Side(style="thin", color="e0d8cc")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_borders(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()


# ─────────────────────────────────────────────────────────────────────────────
# 1. WEEKLY TASKS
# ─────────────────────────────────────────────────────────────────────────────

WEEKLY_TASKS_FULL = {
    1:  {
        "action": ["Take a pregnancy test if not yet confirmed"],
        "research": ["What to expect in early pregnancy", "Foods and drinks to avoid"],
        "consider": ["Who will be the first person you tell?"],
    },
    2:  {
        "action": ["Start folic acid (400mcg daily)", "Stop alcohol and smoking if applicable"],
        "research": ["Folic acid and early neural development", "Pregnancy-safe supplements"],
        "consider": ["When to register with a GP or midwife"],
    },
    3:  {
        "action": ["Stop alcohol, raw fish, soft cheeses, pâté", "Start folic acid if not already"],
        "research": ["Early pregnancy symptoms and what's normal", "NHS Healthy Start scheme"],
        "consider": [],
    },
    4:  {
        "action": ["Book a GP appointment", "Start folic acid (400mcg daily)"],
        "research": ["Vitamin D in pregnancy", "Listeria and food safety"],
        "consider": ["When to tell your employer"],
    },
    5:  {
        "action": ["Register with your GP or midwife", "Note your last menstrual period date"],
        "research": ["How the booking appointment works", "NHS maternity care options"],
        "consider": ["Do you want a midwife-led or consultant-led care pathway?"],
    },
    6:  {
        "action": ["Book your midwife booking appointment (ideally by week 8-10)", "Look into pregnancy multivitamins"],
        "research": ["What happens at the booking appointment", "Pregnancy vitamins — what's needed"],
        "consider": ["Whether to use an NHS midwife, independent midwife, or private care"],
    },
    7:  {
        "action": ["Prepare questions for your booking appointment", "Avoid alcohol, raw fish, soft cheeses, raw meat"],
        "research": ["Hypnobirthing — what it is and whether it interests you", "Mental health support in pregnancy"],
        "consider": ["Who you want to attend scans and appointments with you"],
    },
    8:  {
        "action": ["Attend booking appointment (if booked)", "Complete blood tests requested at booking"],
        "research": ["Down's syndrome, Edwards' syndrome, Patau's screening — how it works", "NIPT (non-invasive prenatal testing) options"],
        "consider": ["Whether to opt into combined screening at 11-13 weeks"],
    },
    9:  {
        "action": ["Rest when you need to — fatigue is normal", "Research maternity pay and leave entitlements"],
        "research": ["SMP (Statutory Maternity Pay) vs occupational maternity pay", "Maternity Allowance if self-employed", "Shared parental leave"],
        "consider": ["When to tell your employer — legal requirement is by week 25"],
    },
    10: {
        "action": ["Book your 12-week dating scan if not yet done", "Think about who you want at your scan"],
        "research": ["What happens at the 12-week scan", "Combined screening results — how they're interpreted"],
        "consider": ["Who will be your primary support person during pregnancy?"],
    },
    11: {
        "action": ["Prepare questions for your sonographer", "Note any symptoms to mention to your midwife"],
        "research": ["Signs of miscarriage and when to seek help", "Ectopic pregnancy symptoms (already past peak risk but worth knowing)"],
        "consider": ["When to share your news more widely"],
    },
    12: {
        "action": ["Attend 12-week dating scan", "Review combined screening results with midwife"],
        "research": ["What 'low risk' and 'higher risk' screening results mean", "NIPT if you want further reassurance"],
        "consider": ["Consider telling close family and friends after the scan"],
    },
    13: {
        "action": ["Book your 20-week anomaly scan", "Consider sharing pregnancy news more widely"],
        "research": ["Second trimester — what to expect physically", "Dental care in pregnancy (free NHS dental treatment)"],
        "consider": ["When to tell your employer (must be by week 25, but earlier is fine)"],
    },
    14: {
        "action": ["Research maternity leave options in detail", "Speak to HR about your company's maternity policy"],
        "research": ["Pram types — travel system vs pushchair vs pram", "Car seat safety — infant carrier vs i-Size"],
        "consider": ["Start a list of big-ticket baby items to budget for"],
    },
    15: {
        "action": ["Start a baby names list if you'd like", "Explore antenatal class options — popular ones book up early"],
        "research": ["NCT classes vs NHS antenatal classes", "Online vs in-person antenatal options"],
        "consider": ["Nursery ideas — no pressure, just early inspiration"],
    },
    16: {
        "action": ["Attend midwife appointment", "Think about your birth preferences — no rush"],
        "research": ["What to expect at the 16-week midwife appointment", "Braxton Hicks vs real contractions"],
        "consider": ["You may start to feel early movements (quickening) — some feel it earlier, some later"],
    },
    17: {
        "action": ["Book antenatal classes if not yet done", "Research hospitals or birth centres nearby"],
        "research": ["Birth centre vs hospital labour ward vs home birth — pros and cons", "Water birth options"],
        "consider": ["Who will be your birth partner — have the conversation now"],
    },
    18: {
        "action": ["Book your 20-week anomaly scan if not yet done", "Start thinking about birth preferences formally"],
        "research": ["Prams and car seats in detail — test in person if possible", "Hypnobirthing courses — book early"],
        "consider": ["Start a hospital bag list (early is fine — you won't pack it yet)"],
    },
    19: {
        "action": ["Prepare questions for your anomaly scan", "Think about hospital bag basics"],
        "research": ["What the anomaly scan checks for", "Placenta previa — what it means if detected"],
        "consider": ["Rest when you can — your body is growing fast now"],
    },
    20: {
        "action": ["Attend 20-week anomaly scan", "Note scan results and any follow-up needed"],
        "research": ["Fetal movements — when to expect them to become regular", "Iron levels in pregnancy — dietary sources"],
        "consider": ["Celebrate — you're halfway there"],
    },
    21: {
        "action": ["Discuss scan results with your midwife", "Start your hospital bag list formally"],
        "research": ["Baby monitors — audio vs video, smart options", "Breastfeeding vs formula — research both without pressure"],
        "consider": ["Think about childcare options — nurseries can have long waiting lists"],
    },
    22: {
        "action": ["Research childcare options and register interest on waiting lists", "Look into your employer's enhanced maternity policy"],
        "research": ["Childcare costs and the 15/30 free hours scheme (eligibility)", "Nursery, childminder, nanny — differences and costs"],
        "consider": ["Research prams — test a few in person if possible"],
    },
    23: {
        "action": ["Book antenatal classes if not yet started", "Start a postpartum plan — meals, support, rest"],
        "research": ["The fourth trimester — what postpartum recovery looks like", "Postpartum mental health — signs of PND"],
        "consider": ["Who will be your birth partner — confirm they understand the role"],
    },
    24: {
        "action": ["Attend midwife appointment", "Review your baby budget and priorities"],
        "research": ["Glucose tolerance test (GTT) — if at risk of gestational diabetes", "Perineal massage — when and how to start"],
        "consider": ["Think about what you actually need for the nursery vs what's nice to have"],
    },
    25: {
        "action": ["Notify employer of pregnancy (legal deadline)", "Submit MATB1 form once received from midwife (usually around 20-25 weeks)"],
        "research": ["MATB1 form — what it is and when you receive it", "SMP eligibility criteria"],
        "consider": ["Draft your birth preferences document — a first attempt, not a final version"],
    },
    26: {
        "action": ["Submit maternity leave paperwork to HR", "Confirm your last working day in writing"],
        "research": ["Shared parental leave — if relevant", "Keeping-in-touch (KIT) days — what they are"],
        "consider": ["Start thinking about a work handover plan"],
    },
    27: {
        "action": ["Book a hospital or birth centre tour", "Start gathering items for your hospital bag"],
        "research": ["Cord blood banking — what it is and costs", "Group B Strep testing — private test available from around 35 weeks"],
        "consider": ["Think about who will look after you postpartum — practical support"],
    },
    28: {
        "action": ["Attend midwife appointment (blood tests this week)", "Review your birth preferences document"],
        "research": ["Anti-D injection (if Rhesus negative)", "Iron supplements if levels low"],
        "consider": ["Sleep may get harder — start prioritising rest now"],
    },
    29: {
        "action": ["Finalise nursery essentials purchases", "Install or plan for baby's car seat"],
        "research": ["Car seat installation — ISOFIX vs belted", "Safe sleep guidelines — The Lullaby Trust"],
        "consider": ["Check your hospital bag list — what do you still need?"],
    },
    30: {
        "action": ["Start packing your hospital bag", "Confirm your birth partner's availability around your due date"],
        "research": ["Signs of labour — show, waters breaking, contractions", "When to call the hospital"],
        "consider": ["Postpartum meals — batch cook or arrange a meal train"],
    },
    31: {
        "action": ["Attend antenatal class if still in session", "Talk through your birth preferences with your midwife"],
        "research": ["Pain relief options in labour — gas and air, epidural, pethidine, TENS", "What a birth plan does and doesn't guarantee"],
        "consider": ["It's okay to slow down — physically and at work"],
    },
    32: {
        "action": ["Add key items to hospital bag", "Pre-register at your hospital or birth centre"],
        "research": ["Newborn care basics — cord care, bathing, feeding cues", "Colostrum harvesting — what it is and whether to try"],
        "consider": ["Discuss any anxieties about birth with your midwife"],
    },
    33: {
        "action": ["Finalise your birth preferences document", "Prepare your home for baby's arrival"],
        "research": ["Caesarean section — what to expect if planned or emergency", "Perineal massage — continue if started"],
        "consider": ["Start thinking about a list of people to notify after birth"],
    },
    34: {
        "action": ["Attend midwife appointment", "Prepare list of people to notify after birth"],
        "research": ["Induction of labour — why it happens and what to expect", "Fetal movements — what's normal at this stage"],
        "consider": ["Hospital bag should be nearly complete"],
    },
    35: {
        "action": ["Check hospital bag is complete", "Know the signs of labour", "Consider private Group B Strep test"],
        "research": ["Group B Strep — risks and testing", "What happens if your waters break at home"],
        "consider": ["Reduce commitments and rest where possible"],
    },
    36: {
        "action": ["Weekly midwife appointments begin from now", "Install car seat and check fitting", "Stock up on postpartum essentials for yourself"],
        "research": ["Postpartum recovery — what your body will need", "Pelvic floor exercises — continue throughout"],
        "consider": ["Prepare a postpartum essentials basket for yourself"],
    },
    37: {
        "action": ["Final check on hospital bag", "Inform your birth partner to stay on call"],
        "research": ["Early labour — what to do and when to go in", "BRAIN acronym for decision-making in labour"],
        "consider": ["Full term — baby could arrive any time now"],
    },
    38: {
        "action": ["Weekly midwife appointment", "Tie up any loose ends at home"],
        "research": ["Overdue pregnancy — monitoring and options", "Membrane sweep — what it involves"],
        "consider": ["Rest as much as possible — this is a legitimate priority"],
    },
    39: {
        "action": ["Stay in contact with your midwife", "Know when to call your midwife or hospital"],
        "research": ["Induction — what to expect if offered", "Positions for labour — upright, on all fours, etc."],
        "consider": ["Walk gently if it helps you feel settled"],
    },
    40: {
        "action": ["Stay in contact with your midwife", "Rest, breathe, trust yourself"],
        "research": ["Post-dates monitoring — what happens after 40 weeks", "Induction timeline if offered"],
        "consider": ["You've done everything you need to do"],
    },
}


def create_weekly_tasks_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Tasks"

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20

    # Header
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "PREGNANCY PLANNER — Weekly Tasks (Exhaustive List for Review)"
    cell.font = Font(color=WHITE, size=13, bold=True)
    cell.fill = PatternFill("solid", fgColor=SAGE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-header
    ws.row_dimensions[2].height = 22
    for col, label in enumerate(["WEEK", "TRIMESTER", "TASK", "TYPE"], start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = label
        style_subheader_row(ws, 2, 4)

    row = 3
    for week in range(1, 41):
        trimester = ("First Trimester" if week <= 12
                     else "Second Trimester" if week <= 26
                     else "Third Trimester")
        data = WEEKLY_TASKS_FULL.get(week, {"action": [], "research": [], "consider": []})
        all_tasks = (
            [(t, "Action") for t in data["action"]] +
            [(t, "Research") for t in data["research"]] +
            [(t, "Consider") for t in data["consider"]]
        )
        if not all_tasks:
            continue

        bg = LIGHT if week % 2 == 0 else None

        for i, (task, task_type) in enumerate(all_tasks):
            ws.row_dimensions[row].height = 18
            ws.cell(row=row, column=1).value = week if i == 0 else ""
            ws.cell(row=row, column=2).value = trimester if i == 0 else ""
            ws.cell(row=row, column=3).value = task
            ws.cell(row=row, column=4).value = task_type

            type_colours = {"Action": "d4edda", "Research": "dbeafe", "Consider": "fef9c3"}
            type_fill = PatternFill("solid", fgColor=type_colours[task_type])
            ws.cell(row=row, column=4).fill = type_fill
            ws.cell(row=row, column=4).font = Font(size=10)

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                if bg and col < 4:
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            row += 1

    apply_borders(ws, 2, row - 1, 1, 4)

    # Freeze top rows
    ws.freeze_panes = "A3"

    wb.save("/Users/annie/pregnancy-planner/weekly_tasks_exhaustive.xlsx")
    print("Saved: weekly_tasks_exhaustive.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 2. HOSPITAL BAG
# ─────────────────────────────────────────────────────────────────────────────

HOSPITAL_BAG = {
    "For You — Labour": [
        "Maternity notes",
        "Birth plan / preferences (printed copy)",
        "Photo ID",
        "Hospital appointment letters",
        "TENS machine (if using)",
        "Birthing ball (if allowed)",
        "Pillow from home",
        "Lip balm",
        "Hair ties / clips",
        "Face mist / cooling spray",
        "Snacks — energy bars, nuts, dried fruit",
        "Drinks — water bottle, isotonic drinks",
        "Phone and charger",
        "Headphones and playlist / hypnobirthing tracks",
        "Camera (optional — phone is fine)",
        "Change for car park",
    ],
    "For You — Postnatal": [
        "Nightgown or pyjamas (x2) — front-opening if breastfeeding",
        "Dressing gown",
        "Slippers (non-slip)",
        "Comfortable high-waisted underwear (x5)",
        "Maternity pads — heavy flow (x2 large packs)",
        "Breast pads (x1 pack)",
        "Nursing bras (x2)",
        "Toiletries bag: toothbrush, toothpaste, face wash, shampoo, deodorant",
        "Dry shampoo",
        "Moisturiser / body lotion",
        "Nipple cream (e.g. Lansinoh)",
        "Comfortable loose clothing to go home in",
        "Flip-flops for the shower",
        "Stool softener / laxatives (ask midwife)",
        "Paracetamol / ibuprofen (hospital may provide, but worth having)",
        "Snacks for after birth",
        "Eye mask and earplugs",
    ],
    "For Baby": [
        "Sleepsuits (x3) — newborn and 0-3 month size",
        "Vests (x3)",
        "Hat (x2)",
        "Scratch mitts (x2 pairs)",
        "Socks (x2 pairs)",
        "Cardigan or jacket",
        "Going-home outfit",
        "Nappies — newborn size (x1 pack)",
        "Cotton wool balls",
        "Water wipes or unscented wipes",
        "Muslin cloths (x6)",
        "Swaddle blanket (x1)",
        "Changing mat (travel)",
        "Nappy bags",
        "Nappy cream (e.g. Bepanthen)",
        "Car seat (installed in car before you leave — essential by law)",
        "Blanket for car / transfer",
    ],
    "Partner / Support Person": [
        "Phone and charger",
        "Change of clothes",
        "Toiletries",
        "Snacks and drinks",
        "Cash and cards",
        "Entertainment for waiting (book, tablet)",
        "List of people to call after birth",
        "Car keys and parking money",
    ],
}


def create_hospital_bag_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hospital Bag"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "HOSPITAL BAG — Exhaustive Checklist for Review"
    cell.font = Font(color=WHITE, size=13, bold=True)
    cell.fill = PatternFill("solid", fgColor=SAGE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for section, items in HOSPITAL_BAG.items():
        # Section header
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = section.upper()
        style_subheader_row(ws, row, 4)
        row += 1

        # Column labels
        for col, label in enumerate(["CATEGORY", "ITEM", "PACKED?", "NOTES"], start=1):
            ws.cell(row=row, column=col).value = label
            ws.cell(row=row, column=col).font = Font(size=9, bold=True, color=HEADER)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="ece6de")
        row += 1

        for i, item in enumerate(items):
            ws.row_dimensions[row].height = 18
            ws.cell(row=row, column=1).value = section
            ws.cell(row=row, column=2).value = item
            ws.cell(row=row, column=3).value = ""  # packed checkbox
            ws.cell(row=row, column=4).value = ""  # notes
            bg = LIGHT if i % 2 == 0 else None
            for col in range(1, 5):
                c = ws.cell(row=row, column=col)
                c.font = Font(size=10)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                if bg:
                    c.fill = PatternFill("solid", fgColor=bg)
            row += 1

        row += 1  # blank gap between sections

    apply_borders(ws, 1, row - 1, 1, 4)
    ws.freeze_panes = "A3"

    wb.save("/Users/annie/pregnancy-planner/hospital_bag_exhaustive.xlsx")
    print("Saved: hospital_bag_exhaustive.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 3. BABY ESSENTIALS
# ─────────────────────────────────────────────────────────────────────────────

BABY_ESSENTIALS = {
    "Sleeping": [
        ("Cot or Moses basket", "Essential", "Safe sleep: firm, flat mattress"),
        ("Mattress — firm and flat", "Essential", "Replace if second-hand"),
        ("Fitted sheets (x3)", "Essential", ""),
        ("Cellular blankets (x2)", "Essential", "Avoid duvets and pillows under 12m"),
        ("Sleeping bags (x2) — appropriate tog", "Essential", ""),
        ("Baby monitor — audio or video", "Recommended", ""),
        ("Blackout blind", "Recommended", ""),
        ("White noise machine", "Optional", ""),
        ("Next-to-me crib (e.g. Snüzpod)", "Optional", "Useful for safe co-sleeping proximity"),
        ("Swaddle wraps (x2)", "Optional", ""),
    ],
    "Feeding — Breastfeeding": [
        ("Nursing pillow (e.g. My Brest Friend, Boppy)", "Recommended", ""),
        ("Nursing bras (x3)", "Recommended", ""),
        ("Breast pads — reusable or disposable", "Essential", ""),
        ("Nipple cream (e.g. Lansinoh)", "Essential", ""),
        ("Breast pump — manual or electric", "Recommended", "Check NHS loan scheme"),
        ("Milk storage bags", "Recommended", ""),
        ("Bottles (x2) even if breastfeeding", "Recommended", "Useful for expressed milk"),
        ("Steriliser — microwave, electric, or cold water", "Recommended", ""),
        ("Muslin cloths (x8+)", "Essential", ""),
        ("Bibs (x6)", "Essential", ""),
    ],
    "Feeding — Formula/Bottle": [
        ("Bottles (x6) — various sizes", "Essential", ""),
        ("Bottle brush", "Essential", ""),
        ("Steriliser", "Essential", ""),
        ("Formula (starter pack)", "Essential", ""),
        ("Perfect Prep machine or kettle + cooling jug", "Optional", ""),
        ("Muslin cloths (x8+)", "Essential", ""),
        ("Bibs (x6)", "Essential", ""),
        ("Bottle warmer", "Optional", ""),
    ],
    "Bathing & Changing": [
        ("Baby bath or bath seat insert", "Recommended", ""),
        ("Hooded towels (x2)", "Essential", ""),
        ("Gentle baby wash and shampoo", "Recommended", "Cotton wool and water for first weeks"),
        ("Cotton wool balls (x2 bags)", "Essential", ""),
        ("Nappy cream (e.g. Bepanthen, Sudocrem)", "Essential", ""),
        ("Changing mat — home use", "Essential", ""),
        ("Changing mat — travel/portable", "Recommended", ""),
        ("Nappy bin (e.g. Tommee Tippee)", "Optional", ""),
        ("Nappies — newborn and size 1 (don't stockpile)", "Essential", ""),
        ("Water wipes or unscented wipes", "Essential", ""),
        ("Nappy bags", "Essential", ""),
        ("Baby nail file or safety clippers", "Essential", ""),
        ("Thermometer — digital rectal or ear", "Essential", ""),
        ("Nasal aspirator (e.g. Frida NoseFrida)", "Recommended", ""),
        ("Baby grooming kit", "Optional", ""),
    ],
    "Clothing (Newborn–3 months)": [
        ("Vests/bodysuits (x7) — short and long sleeve", "Essential", ""),
        ("Sleepsuits/babygrows (x7)", "Essential", ""),
        ("Cardigans (x2)", "Essential", ""),
        ("Hats (x2)", "Essential", ""),
        ("Scratch mitts (x2 pairs)", "Essential", ""),
        ("Socks (x4 pairs)", "Essential", ""),
        ("Snowsuit or warm outer layer (season-dependent)", "Essential", ""),
        ("Going-home outfit", "Recommended", ""),
        ("Muslin bibs (x6)", "Essential", ""),
    ],
    "Out & About": [
        ("Pram / pushchair", "Essential", "Research carefully — major purchase"),
        ("Car seat — infant carrier (Group 0+) or i-Size", "Essential", "Legal requirement"),
        ("Changing bag", "Essential", ""),
        ("Sling or baby carrier", "Optional", "Very useful for hands-free"),
        ("Rain cover for pram", "Essential", "Usually included"),
        ("Sun shade for pram", "Recommended", ""),
        ("Footmuff / cosy toes", "Recommended", ""),
    ],
    "Health & Safety": [
        ("Baby first aid kit", "Recommended", ""),
        ("Digital thermometer", "Essential", ""),
        ("Calpol infant (from 2 months)", "Recommended", "Not needed at birth"),
        ("Stair gates (x2 minimum)", "Essential — before mobile", "Needed ~6-9 months"),
        ("Socket covers", "Recommended", ""),
        ("Baby-safe cupboard locks", "Recommended", "Before mobile"),
        ("Baby monitor", "Recommended", ""),
    ],
    "Nice to Have": [
        ("Baby bouncer chair", "Optional", ""),
        ("Play mat with activity arch", "Optional", "Great for tummy time"),
        ("Baby swing", "Optional", ""),
        ("Door bouncer", "Optional", "From when baby can hold head"),
        ("Jumperoo", "Optional", "From ~4 months"),
        ("Feeding chair / nursing armchair", "Recommended", ""),
    ],
}


def create_baby_essentials_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baby Essentials"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "BABY ESSENTIALS — Exhaustive List for Review"
    cell.font = Font(color=WHITE, size=13, bold=True)
    cell.fill = PatternFill("solid", fgColor=SAGE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for section, items in BABY_ESSENTIALS.items():
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = section.upper()
        style_subheader_row(ws, row, 5)
        row += 1

        for col, label in enumerate(["CATEGORY", "ITEM", "PRIORITY", "NOTES", "GOT IT?"], start=1):
            ws.cell(row=row, column=col).value = label
            ws.cell(row=row, column=col).font = Font(size=9, bold=True, color=HEADER)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="ece6de")
        row += 1

        priority_colours = {
            "Essential": "d4edda",
            "Recommended": "dbeafe",
            "Essential — before mobile": "d4edda",
            "Optional": "fef9c3",
        }

        for i, (item, priority, notes) in enumerate(items):
            ws.row_dimensions[row].height = 18
            ws.cell(row=row, column=1).value = section
            ws.cell(row=row, column=2).value = item
            ws.cell(row=row, column=3).value = priority
            ws.cell(row=row, column=4).value = notes
            ws.cell(row=row, column=5).value = ""

            bg = LIGHT if i % 2 == 0 else None
            p_fill = PatternFill("solid", fgColor=priority_colours.get(priority, "FFFFFF"))
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                c.font = Font(size=10)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                if col == 3:
                    c.fill = p_fill
                elif bg:
                    c.fill = PatternFill("solid", fgColor=bg)
            row += 1

        row += 1

    apply_borders(ws, 1, row - 1, 1, 5)
    ws.freeze_panes = "A3"

    wb.save("/Users/annie/pregnancy-planner/baby_essentials_exhaustive.xlsx")
    print("Saved: baby_essentials_exhaustive.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 4. WORK & ADMIN
# ─────────────────────────────────────────────────────────────────────────────

WORK_ADMIN = {
    "During Pregnancy — Work & HR": [
        ("Notify employer of pregnancy in writing", "By week 25 (legal)", "Earlier is fine and often better"),
        ("Request a workplace risk assessment", "As soon as possible", "Your employer is legally required to do this"),
        ("Confirm your maternity leave start date with HR", "By week 25", ""),
        ("Submit MATB1 certificate to employer", "After 20 weeks, once received from midwife", ""),
        ("Check company maternity pay policy vs SMP", "Early", ""),
        ("Confirm enhanced maternity pay entitlement (if applicable)", "Early", ""),
        ("Review your employment contract re: maternity rights", "Early", ""),
        ("Discuss keeping-in-touch (KIT) days with employer", "Before leave", "Up to 10 KIT days allowed"),
        ("Discuss shared parental leave with partner if relevant", "Before week 25", ""),
        ("Plan work handover and document key processes", "Before leave", ""),
        ("Brief your manager and team on handover", "Before leave", ""),
        ("Set up out-of-office / handover email", "Day before leave", ""),
        ("Arrange laptop / access return if needed", "Before leave", ""),
    ],
    "During Pregnancy — Benefits & Finance": [
        ("Apply for Healthy Start vouchers (if eligible)", "As early as possible", "Free vitamins and food vouchers"),
        ("Check NHS entitlements — free prescriptions and dental care", "Immediately", "Apply for maternity exemption certificate (FW8 form)"),
        ("Apply for maternity exemption certificate (FW8)", "Early pregnancy", "Gives free NHS prescriptions and dental"),
        ("Apply for Maternity Allowance (if self-employed or insufficient SMP)", "From week 26", ""),
        ("Research Child Benefit — apply after birth", "Before birth", ""),
        ("Check eligibility for Universal Credit / tax credits", "Early", ""),
        ("Review household budget for mat leave income reduction", "Second trimester", ""),
        ("Check life insurance and income protection cover", "Any time", ""),
        ("Update pension contributions / nominations if needed", "Any time", ""),
        ("Update your will", "Before birth", "Especially to name guardian"),
    ],
    "Before Baby Arrives — Practical": [
        ("Register baby's name choice (decide before birth if possible)", "Pre-birth", ""),
        ("Research GP registration for baby (at your existing practice if possible)", "Before birth", ""),
        ("Set up a nursery / sleeping space", "By week 36", ""),
        ("Install car seat and test fit", "By week 36", ""),
        ("Prepare hospital bag", "By week 35", ""),
        ("Write birth preferences / plan", "By week 34", ""),
        ("Arrange pet / childcare cover for during labour", "Before due date", ""),
        ("Notify close family / friends of due date and contact plan", "Before due date", ""),
        ("Prepare postpartum essentials basket", "Before birth", ""),
        ("Batch cook and freeze meals", "Weeks 34-38", ""),
        ("Set up a postnatal support plan (visitors, help, rest)", "Before birth", ""),
    ],
    "After Birth — Immediate (0–4 weeks)": [
        ("Register the birth", "Within 42 days (England/Wales), 21 days (Scotland)", "At local register office"),
        ("Apply for Child Benefit", "As soon as possible after birth", "Backdated up to 3 months"),
        ("Notify HMRC of new child (for tax credits if applicable)", "Within 1 month", ""),
        ("Add baby to your GP practice", "First week", ""),
        ("Check in with health visitor", "They will contact you", ""),
        ("Apply for child's passport (if travel planned)", "When ready", ""),
        ("Open a Junior ISA or savings account", "Any time", ""),
        ("Update life insurance / write-in trust if needed", "After birth", ""),
        ("Notify your employer of baby's birth date (to confirm mat leave)", "Within 2 weeks", ""),
    ],
    "After Birth — Longer Term (1–6 months)": [
        ("Notify employer of return to work date (if different from original)", "8 weeks before return", ""),
        ("Research childcare and register (if not done)", "As early as possible — long waitlists", ""),
        ("Apply for 15/30 free childcare hours (eligible from term after 9m)", "Before return to work", ""),
        ("Apply for childcare vouchers or Tax-Free Childcare scheme", "Before return to work", ""),
        ("Review maternity pay final date and return-to-work date", "Before return", ""),
        ("Update will now that baby is here", "Within first year", ""),
        ("Check pensions / beneficiaries updated", "Within first year", ""),
        ("Update home/car insurance if needed", "Any time", ""),
    ],
}


def create_admin_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work & Admin"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "WORK & ADMIN CHECKLIST — Exhaustive List for Review"
    cell.font = Font(color=WHITE, size=13, bold=True)
    cell.fill = PatternFill("solid", fgColor=SAGE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for section, items in WORK_ADMIN.items():
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = section.upper()
        style_subheader_row(ws, row, 5)
        row += 1

        for col, label in enumerate(["CATEGORY", "TASK", "WHEN", "NOTES", "DONE?"], start=1):
            ws.cell(row=row, column=col).value = label
            ws.cell(row=row, column=col).font = Font(size=9, bold=True, color=HEADER)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="ece6de")
        row += 1

        for i, (task, when, notes) in enumerate(items):
            ws.row_dimensions[row].height = 18
            ws.cell(row=row, column=1).value = section
            ws.cell(row=row, column=2).value = task
            ws.cell(row=row, column=3).value = when
            ws.cell(row=row, column=4).value = notes
            ws.cell(row=row, column=5).value = ""

            bg = LIGHT if i % 2 == 0 else None
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                c.font = Font(size=10)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                if bg:
                    c.fill = PatternFill("solid", fgColor=bg)
            row += 1

        row += 1

    apply_borders(ws, 1, row - 1, 1, 5)
    ws.freeze_panes = "A3"

    wb.save("/Users/annie/pregnancy-planner/work_admin_exhaustive.xlsx")
    print("Saved: work_admin_exhaustive.xlsx")


if __name__ == "__main__":
    create_weekly_tasks_excel()
    create_hospital_bag_excel()
    create_baby_essentials_excel()
    create_admin_excel()
    print("\nAll 4 Excel files created.")
